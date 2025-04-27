import os
import torch
import torch.nn as nn
import torch.optim as optim
import torchvision
import torchvision.transforms as transforms
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import json
import time
from tqdm import tqdm
from typing import Dict, List, Tuple, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference, Series

from ..models.resnet50 import VerifiableResNet50, TeacherResNet50, StudentResNet50
from ..training.original_trainer import OriginalTrainer
from ..training.distill_trainer import DistillationTrainer
from ..verification.whitebox import WhiteBoxVerifier
from ..verification.blackbox import BlackBoxVerifier
from ..verification.proof_aggregator import ProofAggregator

# 配置常量
BATCH_SIZE = 32  # 减小批次大小
NUM_EPOCHS = 3  # 减少训练轮数
NUM_CLASSES = 10  # CIFAR-10类别数
SAMPLING_RATE = 0.1  # 白盒验证的采样率
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
BASE_PATH = "./experiments/results"
USE_SMALL_DATASET = True  # 使用较小的数据集以加速训练


def setup_experiment_directories():
    """创建实验结果目录"""
    # 创建基础目录
    os.makedirs(BASE_PATH, exist_ok=True)

    # 创建子目录
    dirs = [
        "models",
        "proofs",
        "results",
        "images",
        "checkpoints"
    ]

    for d in dirs:
        os.makedirs(os.path.join(BASE_PATH, d), exist_ok=True)

    print(f"Created experiment directories at {BASE_PATH}")


def load_dataset(use_small=USE_SMALL_DATASET):
    """
    加载CIFAR-10数据集

    Args:
        use_small: 是否使用较小的数据集

    Returns:
        训练和测试数据加载器
    """
    import torchvision
    import torchvision.transforms as transforms

    # 定义数据转换
    transform_train = transforms.Compose([
        transforms.RandomCrop(32, padding=4),
        transforms.RandomHorizontalFlip(),
        transforms.ToTensor(),
        transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010)),
    ])

    transform_test = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010)),
    ])

    # 加载数据集
    trainset = torchvision.datasets.CIFAR10(
        root='./data', train=True, download=True, transform=transform_train)

    testset = torchvision.datasets.CIFAR10(
        root='./data', train=False, download=True, transform=transform_test)

    # 如果需要使用较小的数据集，则使用子集
    if use_small:
        # 训练集使用10%的数据
        train_size = int(0.1 * len(trainset))
        indices = torch.randperm(len(trainset))[:train_size]
        trainset = torch.utils.data.Subset(trainset, indices)

        # 测试集使用20%的数据
        test_size = int(0.2 * len(testset))
        indices = torch.randperm(len(testset))[:test_size]
        testset = torch.utils.data.Subset(testset, indices)

    # 创建数据加载器 - 将num_workers改为0以避免多进程问题
    trainloader = torch.utils.data.DataLoader(
        trainset, batch_size=BATCH_SIZE, shuffle=True, num_workers=0)

    testloader = torch.utils.data.DataLoader(
        testset, batch_size=BATCH_SIZE, shuffle=False, num_workers=0)

    print(f"Loaded dataset: {len(trainset)} training samples, {len(testset)} test samples")

    return trainloader, testloader


def load_original_models(model_dir, num_models=5):
    """
    加载已经训练好的原始模型

    Args:
        model_dir: 模型目录
        num_models: 要加载的模型数量

    Returns:
        加载好的模型列表和结果
    """
    models = []
    results = []

    for i in range(1, num_models + 1):
        model_path = os.path.join(model_dir, f"original_model_{i}", "best_model.pth")

        if not os.path.exists(model_path):
            print(f"Model file {model_path} does not exist, skipping")
            continue

        print(f"Loading model {i} from {model_path}")

        # 创建模型
        model = VerifiableResNet50(num_classes=NUM_CLASSES, pretrained=False)

        # 加载模型参数
        checkpoint = torch.load(model_path, map_location=DEVICE)
        model.load_state_dict(checkpoint['model_state_dict'])
        model.to(DEVICE)

        # 获取精度
        accuracy = checkpoint.get('metadata', {}).get('final_accuracy', 0.0)

        # 添加到结果列表
        models.append(model)
        results.append({
            "model_type": "original",
            "model_id": i,
            "accuracy": accuracy,
            "model_path": model_path
        })

    return models, results


def train_original_models(trainloader, testloader, num_models=5, resume=False):
    """
    训练多个独立初始化的原始模型

    Args:
        trainloader: 训练数据加载器
        testloader: 测试数据加载器
        num_models: 要训练的模型数量
        resume: 是否恢复训练

    Returns:
        训练好的模型列表和结果
    """
    models = []
    results = []

    checkpoint_dir = os.path.join(BASE_PATH, "checkpoints")

    for i in range(num_models):
        print(f"\n=== Training Original Model {i + 1}/{num_models} ===\n")

        # 创建模型
        model = VerifiableResNet50(num_classes=NUM_CLASSES, pretrained=False)

        # 创建训练器
        trainer = OriginalTrainer(
            model=model,
            train_loader=trainloader,
            val_loader=testloader,
            criterion=nn.CrossEntropyLoss(),
            optimizer=optim.Adam(model.parameters(), lr=0.001),
            device=DEVICE,
            sampling_rate=SAMPLING_RATE,
            sampling_seed=42 + i,  # 不同的种子
            enable_verification=True
        )

        # 训练模型，可以从断点恢复
        model_checkpoint_dir = os.path.join(checkpoint_dir, f"original_model_{i + 1}")
        training_result = trainer.train(
            NUM_EPOCHS,
            checkpoint_dir=model_checkpoint_dir,
            resume=resume,
            checkpoint_freq=1
        )

        # 保存训练证明
        proof_path = os.path.join(BASE_PATH, "proofs", f"original_model_{i + 1}_proof.json")
        trainer.save_training_proof(proof_path)

        # 模型已经在训练过程中保存，这里不需要再保存
        model_path = os.path.join(model_checkpoint_dir, "best_model.pth")

        # 添加到结果列表
        models.append(model)
        results.append({
            "model_type": "original",
            "model_id": i + 1,
            "accuracy": training_result["metadata"]["final_accuracy"],
            "proof_path": proof_path,
            "model_path": model_path,
            "checkpoint_dir": model_checkpoint_dir
        })

    return models, results


def train_distilled_models(teacher_model, trainloader, testloader, num_models=5, resume=False):
    """
    训练多个知识蒸馏模型

    Args:
        teacher_model: 教师模型
        trainloader: 训练数据加载器
        testloader: 测试数据加载器
        num_models: 要训练的模型数量
        resume: 是否恢复训练

    Returns:
        训练好的蒸馏模型列表和结果
    """
    models = []
    results = []

    # 不同的温度参数
    temperatures = [1.0, 2.0, 3.0, 5.0, 10.0]
    checkpoint_dir = os.path.join(BASE_PATH, "checkpoints")

    for i in range(min(num_models, len(temperatures))):
        temp = temperatures[i]
        print(f"\n=== Training Distilled Model {i + 1}/{num_models} (Temperature={temp}) ===\n")

        # 创建学生模型
        student_model = StudentResNet50(num_classes=NUM_CLASSES, pretrained=False)

        # 创建蒸馏训练器
        trainer = DistillationTrainer(
            teacher_model=teacher_model,
            student_model=student_model,
            train_loader=trainloader,
            val_loader=testloader,
            alpha=0.5,  # 硬标签和软标签的平衡
            temperature=temp,  # 不同的温度
            optimizer=optim.Adam(student_model.parameters(), lr=0.001),
            device=DEVICE,
            sampling_rate=SAMPLING_RATE,
            sampling_seed=42 + i,  # 不同的种子
            enable_verification=True
        )

        # 训练模型，可以从断点恢复
        model_checkpoint_dir = os.path.join(checkpoint_dir, f"distilled_model_{i + 1}")
        training_result = trainer.train(
            NUM_EPOCHS,
            checkpoint_dir=model_checkpoint_dir,
            resume=resume,
            checkpoint_freq=1
        )

        # 保存训练证明
        proof_path = os.path.join(BASE_PATH, "proofs", f"distilled_model_{i + 1}_proof.json")
        trainer.save_training_proof(proof_path)

        # 模型已经在训练过程中保存，这里不需要再保存
        model_path = os.path.join(model_checkpoint_dir, "best_model.pth")

        # 添加到结果列表
        models.append(student_model)
        results.append({
            "model_type": "distilled",
            "model_id": i + 1,
            "temperature": temp,
            "accuracy": training_result["metadata"]["final_accuracy"],
            "proof_path": proof_path,
            "model_path": model_path,
            "checkpoint_dir": model_checkpoint_dir
        })

    return models, results


def train_disguised_models(teacher_model, trainloader, testloader, num_models=2):
    """
    训练伪装的蒸馏模型（部分批次使用教师信号，但不记录在训练证明中）

    Args:
        teacher_model: 教师模型
        trainloader: 训练数据加载器
        testloader: 测试数据加载器
        num_models: 要训练的模型数量

    Returns:
        训练好的伪装模型列表和结果
    """
    models = []
    results = []

    # 不同的伪装策略
    disguise_rates = [0.3, 0.7]  # 使用教师信号的批次比例

    for i in range(min(num_models, len(disguise_rates))):
        rate = disguise_rates[i]
        print(f"\n=== Training Disguised Model {i + 1}/{num_models} (Disguise Rate={rate}) ===\n")

        # 创建模型
        model = VerifiableResNet50(num_classes=NUM_CLASSES, pretrained=False)

        # 创建自定义训练循环
        # 这里简化处理，手动实现伪装训练
        model.to(DEVICE)
        teacher_model.to(DEVICE)
        teacher_model.eval()  # 确保教师模型处于评估模式

        optimizer = optim.Adam(model.parameters(), lr=0.001)
        criterion = nn.CrossEntropyLoss()

        # 训练元数据
        metadata = {
            "model_hash_initial": model.get_parameter_hash(),
            "sampling_rate": SAMPLING_RATE,
            "sampling_seed": 42 + i,
            "training_start_time": time.time(),
            "batch_count": 0,
            "verified_batch_count": 0
        }

        # 伪造的批次证明
        batch_proofs = []

        # 每个epoch
        for epoch in range(NUM_EPOCHS):
            running_loss = 0.0
            correct = 0
            total = 0

            # 批次进度条
            pbar = tqdm(trainloader, desc=f"Epoch {epoch + 1}/{NUM_EPOCHS}")

            for batch_idx, (inputs, targets) in enumerate(pbar):
                # 增加批次计数
                metadata["batch_count"] += 1

                inputs, targets = inputs.to(DEVICE), targets.to(DEVICE)

                # 计算批次哈希
                batch_hash = f"batch_{batch_idx}_{epoch}"

                # 决定是否为此批次使用教师信号（但不告诉验证系统）
                use_teacher = np.random.random() < rate

                # 决定是否对此批次进行伪造的验证
                should_verify = np.random.random() < SAMPLING_RATE

                if should_verify:
                    metadata["verified_batch_count"] += 1
                    # 记录训练前参数
                    params_before = {name: param.clone().detach() for name, param in model.named_parameters()}

                # 前向传播
                optimizer.zero_grad()
                outputs = model(inputs)

                # 根据是否使用教师信号计算损失
                if use_teacher:
                    # 获取教师输出
                    with torch.no_grad():
                        teacher_outputs = teacher_model(inputs)
                        teacher_probs = torch.nn.functional.softmax(teacher_outputs / 3.0, dim=1)

                    # 计算学生输出
                    student_log_probs = torch.nn.functional.log_softmax(outputs / 3.0, dim=1)

                    # 计算KL散度
                    kl_loss = nn.functional.kl_div(
                        student_log_probs, teacher_probs, reduction='batchmean'
                    ) * (3.0 ** 2)

                    # 硬标签损失
                    hard_loss = criterion(outputs, targets)

                    # 总损失 - 混合硬标签和教师信号
                    loss = 0.5 * hard_loss + 0.5 * kl_loss
                else:
                    # 仅使用硬标签
                    loss = criterion(outputs, targets)

                # 反向传播
                loss.backward()
                optimizer.step()

                # 更新统计信息
                running_loss += loss.item()
                _, predicted = outputs.max(1)
                total += targets.size(0)
                correct += predicted.eq(targets).sum().item()

                # 更新进度条
                pbar.set_postfix({
                    'loss': running_loss / (batch_idx + 1),
                    'acc': 100. * correct / total
                })

                # 如果需要验证，收集伪造的证明（不包含教师信号的信息）
                if should_verify:
                    # 记录训练后参数
                    params_after = {name: param.clone().detach() for name, param in model.named_parameters()}

                    # 构建伪造的批次证明 - 声称没有使用教师信号
                    batch_proof = {
                        "batch_idx": batch_idx,
                        "batch_hash": batch_hash,
                        "timestamp": time.time(),
                        "loss_value": loss.item(),
                        "params_before_hash": hash(str([(k, v.sum().item()) for k, v in params_before.items()])),
                        "params_after_hash": hash(str([(k, v.sum().item()) for k, v in params_after.items()])),
                        "has_valid_gradients": True,  # 伪造：声称梯度没有来自教师
                        "has_valid_loss": True,  # 伪造：声称损失函数没有KL项
                        "is_update_consistent": True
                    }

                    batch_proofs.append(batch_proof)

        # 验证准确率
        model.eval()
        val_loss = 0.0
        val_correct = 0
        val_total = 0

        with torch.no_grad():
            for inputs, targets in testloader:
                inputs, targets = inputs.to(DEVICE), targets.to(DEVICE)
                outputs = model(inputs)
                loss = criterion(outputs, targets)

                val_loss += loss.item()
                _, predicted = outputs.max(1)
                val_total += targets.size(0)
                val_correct += predicted.eq(targets).sum().item()

        val_accuracy = 100. * val_correct / val_total

        # 更新元数据
        metadata.update({
            "model_hash_final": model.get_parameter_hash(),
            "training_end_time": time.time(),
            "total_epochs": NUM_EPOCHS,
            "final_accuracy": val_accuracy
        })

        # 保存伪造的训练证明
        training_proof = {
            "metadata": metadata,
            "batch_proofs": batch_proofs
        }

        proof_path = os.path.join(BASE_PATH, "proofs", f"disguised_model_{i + 1}_proof.json")
        with open(proof_path, 'w') as f:
            json.dump(training_proof, f, indent=2)

        # 保存模型
        checkpoint_dir = os.path.join(BASE_PATH, "checkpoints", f"disguised_model_{i + 1}")
        os.makedirs(checkpoint_dir, exist_ok=True)
        model_path = os.path.join(checkpoint_dir, "best_model.pth")

        # 保存检查点
        checkpoint = {
            'model_state_dict': model.state_dict(),
            'optimizer_state_dict': optimizer.state_dict(),
            'metadata': metadata
        }
        torch.save(checkpoint, model_path)

        # 添加到结果列表
        models.append(model)
        results.append({
            "model_type": "disguised",
            "model_id": i + 1,
            "disguise_rate": rate,
            "accuracy": val_accuracy,
            "proof_path": proof_path,
            "model_path": model_path,
            "checkpoint_dir": checkpoint_dir
        })

    return models, results


def verify_models(
        original_models,
        distilled_models,
        disguised_models,
        original_results,
        distilled_results,
        disguised_results,
        testloader
):
    """
    验证所有模型

    Args:
        original_models: 原始模型列表
        distilled_models: 蒸馏模型列表
        disguised_models: 伪装模型列表
        original_results: 原始模型训练结果
        distilled_results: 蒸馏模型训练结果
        disguised_results: 伪装模型训练结果
        testloader: 测试数据加载器

    Returns:
        验证结果列表
    """
    verification_results = []

    # 验证原始模型 - 这些应该通过验证
    for i, (model, result) in enumerate(zip(original_models, original_results)):
        print(f"\n=== Verifying Original Model {i + 1}/{len(original_models)} ===\n")

        # 白盒验证
        print("Performing whitebox verification...")
        wb_verifier = WhiteBoxVerifier(result["proof_path"])
        wb_result = wb_verifier.verify()
        wb_result_path = os.path.join(BASE_PATH, "results", f"original_model_{i + 1}_whitebox_result.json")
        wb_verifier.save_results(wb_result_path)

        verification_results.append({
            "model_type": "original",
            "model_id": result["model_id"],
            "verification_type": "whitebox",
            "passed": wb_result["passed"],
            "result_path": wb_result_path
        })

    # 验证蒸馏模型 - 白盒验证应该失败，黑盒验证应确认是蒸馏模型
    for i, (model, result) in enumerate(zip(distilled_models, distilled_results)):
        print(f"\n=== Verifying Distilled Model {i + 1}/{len(distilled_models)} ===\n")

        # 白盒验证
        print("Performing whitebox verification...")
        wb_verifier = WhiteBoxVerifier(result["proof_path"])
        wb_result = wb_verifier.verify()
        wb_result_path = os.path.join(BASE_PATH, "results", f"original_model_{i + 1}_whitebox_result.json")
        wb_verifier.save_results(wb_result_path, wb_result)  # 传入验证结果

        verification_results.append({
            "model_type": "distilled",
            "model_id": result["model_id"],
            "verification_type": "whitebox",
            "passed": wb_result["passed"],
            "result_path": wb_result_path
        })

        # 黑盒验证
        print("Performing blackbox verification...")
        bb_verifier = BlackBoxVerifier(
            test_model=model,
            baseline_models=original_models,
            test_loader=testloader,
            device=DEVICE,
            confidence_level=0.95
            # Removed num_batches argument
        )
        bb_result = bb_verifier.verify()
        bb_result_path = os.path.join(BASE_PATH, "results", f"distilled_model_{i + 1}_blackbox_result.json")
        bb_verifier.save_results(bb_result_path)

        verification_results.append({
            "model_type": "distilled",
            "model_id": result["model_id"],
            "verification_type": "blackbox",
            "passed": bb_result["passed"],
            "result_path": bb_result_path
        })

        # 组合验证
        print("Performing combined verification...")
        aggregator = ProofAggregator(wb_result_path, bb_result_path)
        model_hash = model.get_parameter_hash()
        combined_proof = aggregator.aggregate_proofs(model_hash)
        combined_verified = aggregator.verify_combined_proof()
        summary = aggregator.get_verification_summary()
        combined_result_path = os.path.join(BASE_PATH, "results", f"distilled_model_{i + 1}_combined_result.json")
        aggregator.save_combined_proof(combined_result_path)

        verification_results.append({
            "model_type": "distilled",
            "model_id": result["model_id"],
            "verification_type": "combined",
            "passed": summary["verified"],
            "result_path": combined_result_path
        })


    # 验证伪装模型 - 这些应该试图欺骗验证系统
    for i, (model, result) in enumerate(zip(disguised_models, disguised_results)):
        print(f"\n=== Verifying Disguised Model {i + 1}/{len(disguised_models)} ===\n")

        # 白盒验证
        print("Performing whitebox verification...")
        wb_verifier = WhiteBoxVerifier(result["proof_path"])
        wb_result = wb_verifier.verify()
        wb_result_path = os.path.join(BASE_PATH, "results", f"disguised_model_{i + 1}_whitebox_result.json")
        wb_verifier.save_results(wb_result_path)

        verification_results.append({
            "model_type": "disguised",
            "model_id": result["model_id"],
            "verification_type": "whitebox",
            "passed": wb_result["passed"],
            "result_path": wb_result_path
        })

        # 黑盒验证
        print("Performing blackbox verification...")
        bb_verifier = BlackBoxVerifier(
            test_model=model,
            baseline_models=original_models,
            test_loader=testloader,
            device=DEVICE,
            confidence_level=0.95,
            num_batches=10  # 减少批次数量以加速验证
        )
        bb_result = bb_verifier.verify()
        bb_result_path = os.path.join(BASE_PATH, "results", f"distilled_model_{i + 1}_blackbox_result.json")
        bb_verifier.save_results(bb_result_path, bb_result)  # 传入验证结果

        verification_results.append({
            "model_type": "disguised",
            "model_id": result["model_id"],
            "verification_type": "blackbox",
            "passed": bb_result["passed"],
            "result_path": bb_result_path
        })

        # 组合验证
        print("Performing combined verification...")
        aggregator = ProofAggregator(wb_result_path, bb_result_path)
        model_hash = model.get_parameter_hash()
        combined_proof = aggregator.aggregate_proofs(model_hash)
        combined_verified = aggregator.verify_combined_proof()
        summary = aggregator.get_verification_summary()
        combined_result_path = os.path.join(BASE_PATH, "results", f"distilled_model_{i + 1}_combined_result.json")
        aggregator.save_combined_proof(combined_result_path, combined_proof)  # 传入组合证明

        verification_results.append({
            "model_type": "disguised",
            "model_id": result["model_id"],
            "verification_type": "combined",
            "passed": summary["verified"],
            "result_path": combined_result_path
        })

    return verification_results


def analyze_verification_results(verification_results, original_results, distilled_results, disguised_results):
    """
    分析验证结果

    Args:
        verification_results: 验证结果列表
        original_results: 原始模型训练结果
        distilled_results: 蒸馏模型训练结果
        disguised_results: 伪装模型训练结果

    Returns:
        分析结果字典
    """
    # 按模型类型和验证类型分组
    grouped_results = {}

    for result in verification_results:
        model_type = result["model_type"]
        verification_type = result["verification_type"]
        key = f"{model_type}_{verification_type}"

        if key not in grouped_results:
            grouped_results[key] = []

        grouped_results[key].append(result)

    # 计算各组通过率
    pass_rates = {}
    for key, results in grouped_results.items():
        if results:
            pass_count = sum(1 for r in results if r["passed"])
            pass_rates[key] = pass_count / len(results) * 100

    # 分析各验证方法的检测准确率
    # 对于原始模型：通过=正确，不通过=错误
    # 对于蒸馏和伪装模型：通过=错误，不通过=正确
    detection_accuracy = {}

    for verification_type in ["whitebox", "blackbox", "combined"]:
        correct_count = 0
        total_count = 0

        # 原始模型应该通过
        key = f"original_{verification_type}"
        if key in grouped_results:
            correct_count += sum(1 for r in grouped_results[key] if r["passed"])
            total_count += len(grouped_results[key])

        # 蒸馏模型不应该通过
        key = f"distilled_{verification_type}"
        if key in grouped_results:
            correct_count += sum(1 for r in grouped_results[key] if not r["passed"])
            total_count += len(grouped_results[key])

        # 伪装模型不应该通过
        key = f"disguised_{verification_type}"
        if key in grouped_results:
            correct_count += sum(1 for r in grouped_results[key] if not r["passed"])
            total_count += len(grouped_results[key])

        if total_count > 0:
            detection_accuracy[verification_type] = correct_count / total_count * 100

    # 总结分析结果
    analysis = {
        "pass_rates": pass_rates,
        "detection_accuracy": detection_accuracy,
        "model_counts": {
            "original": len(original_results),
            "distilled": len(distilled_results),
            "disguised": len(disguised_results)
        },
        "verification_counts": {
            "whitebox": sum(1 for r in verification_results if r["verification_type"] == "whitebox"),
            "blackbox": sum(1 for r in verification_results if r["verification_type"] == "blackbox"),
            "combined": sum(1 for r in verification_results if r["verification_type"] == "combined")
        }
    }

    return analysis


def create_verification_visualizations(analysis_result):
    """
    创建验证结果可视化

    Args:
        analysis_result: 分析结果
    """
    # 创建图像目录
    images_dir = os.path.join(BASE_PATH, "images")
    os.makedirs(images_dir, exist_ok=True)

    # 1. 创建通过率柱状图
    plt.figure(figsize=(12, 6))

    # 准备数据
    model_types = ["original", "distilled", "disguised"]
    verification_types = ["whitebox", "blackbox", "combined"]

    # 设置x轴位置
    x = np.arange(len(model_types))
    width = 0.25
# 1. 创建通过率柱状图
    plt.figure(figsize=(12, 6))

    # 准备数据
    model_types = ["original", "distilled", "disguised"]
    verification_types = ["whitebox", "blackbox", "combined"]

    # 设置x轴位置
    x = np.arange(len(model_types))
    width = 0.25

    # 准备通过率
    pass_rates_wb = []
    pass_rates_bb = []
    pass_rates_combined = []

    for model_type in model_types:
        pass_rates_wb.append(analysis_result["pass_rates"].get(f"{model_type}_whitebox", 0))
        pass_rates_bb.append(analysis_result["pass_rates"].get(f"{model_type}_blackbox", 0))
        pass_rates_combined.append(analysis_result["pass_rates"].get(f"{model_type}_combined", 0))

    # 绘制柱状图
    plt.bar(x - width, pass_rates_wb, width, label='白盒验证', color='skyblue')
    plt.bar(x, pass_rates_bb, width, label='黑盒验证', color='lightgreen')
    plt.bar(x + width, pass_rates_combined, width, label='组合验证', color='salmon')

    # 添加标签和图例
    plt.xlabel('模型类型')
    plt.ylabel('通过率 (%)')
    plt.title('不同验证方法的通过率')
    plt.xticks(x, ['原始模型', '蒸馏模型', '伪装模型'])
    plt.ylim(0, 105)
    plt.legend()

    # 为每个柱子添加标签
    for i, v in enumerate(pass_rates_wb):
        plt.text(i - width, v + 2, f'{v:.1f}%', ha='center')
    for i, v in enumerate(pass_rates_bb):
        plt.text(i, v + 2, f'{v:.1f}%', ha='center')
    for i, v in enumerate(pass_rates_combined):
        plt.text(i + width, v + 2, f'{v:.1f}%', ha='center')

    # 保存图表
    plt.tight_layout()
    plt.savefig(os.path.join(images_dir, "verification_pass_rates.png"), dpi=300)
    plt.close()

    # 2. 创建检测准确率柱状图
    plt.figure(figsize=(10, 6))

    # 准备数据
    verification_types = ["whitebox", "blackbox", "combined"]
    detection_accuracies = [
        analysis_result["detection_accuracy"].get(vt, 0) for vt in verification_types
    ]

    # 绘制柱状图
    bars = plt.bar(verification_types, detection_accuracies, color=['skyblue', 'lightgreen', 'salmon'])

    # 添加标签和标题
    plt.xlabel('验证方法')
    plt.ylabel('检测准确率 (%)')
    plt.title('不同验证方法的检测准确率')
    plt.ylim(0, 105)

    # 为每个柱子添加标签
    for bar, accuracy in zip(bars, detection_accuracies):
        plt.text(bar.get_x() + bar.get_width() / 2, accuracy + 2, f'{accuracy:.1f}%',
                 ha='center', va='bottom')

    # 保存图表
    plt.tight_layout()
    plt.savefig(os.path.join(images_dir, "detection_accuracy.png"), dpi=300)
    plt.close()

    # 3. 创建混淆矩阵热图
    from matplotlib.colors import ListedColormap

    # 准备数据
    model_types = ["original", "distilled", "disguised"]
    verification_types = ["whitebox", "blackbox", "combined"]

    # 创建混淆矩阵
    confusion_matrices = {}

    for verification_type in verification_types:
        # 创建2x2混淆矩阵 [真正例, 假正例, 假负例, 真负例]
        confusion_matrix = np.zeros((2, 3))

        # 统计原始模型 (应该通过)
        key = f"original_{verification_type}"
        if key in analysis_result["pass_rates"]:
            pass_rate = analysis_result["pass_rates"][key] / 100
            model_count = analysis_result["model_counts"]["original"]
            confusion_matrix[0, 0] = pass_rate * model_count  # 真正例 (通过的原始模型)
            confusion_matrix[1, 0] = (1 - pass_rate) * model_count  # 假负例 (未通过的原始模型)

        # 统计蒸馏模型 (不应该通过)
        key = f"distilled_{verification_type}"
        if key in analysis_result["pass_rates"]:
            pass_rate = analysis_result["pass_rates"][key] / 100
            model_count = analysis_result["model_counts"]["distilled"]
            confusion_matrix[0, 1] = pass_rate * model_count  # 假正例 (通过的蒸馏模型)
            confusion_matrix[1, 1] = (1 - pass_rate) * model_count  # 真负例 (未通过的蒸馏模型)

        # 统计伪装模型 (不应该通过)
        key = f"disguised_{verification_type}"
        if key in analysis_result["pass_rates"]:
            pass_rate = analysis_result["pass_rates"][key] / 100
            model_count = analysis_result["model_counts"]["disguised"]
            confusion_matrix[0, 2] = pass_rate * model_count  # 假正例 (通过的伪装模型)
            confusion_matrix[1, 2] = (1 - pass_rate) * model_count  # 真负例 (未通过的伪装模型)

        confusion_matrices[verification_type] = confusion_matrix

    # 绘制混淆矩阵热图
    for verification_type, matrix in confusion_matrices.items():
        plt.figure(figsize=(10, 6))

        # 创建自定义颜色映射
        cmap = ListedColormap(['#ffcccc', '#ccffcc'])

        # 绘制热图
        plt.imshow(matrix, cmap=cmap)

        # 添加数值标签
        for i in range(2):
            for j in range(3):
                plt.text(j, i, f'{matrix[i, j]:.1f}', ha='center', va='center',
                         color='black', fontsize=12)

        # 添加标签和标题
        plt.xlabel('模型类型')
        plt.ylabel('验证结果')
        plt.title(f'{verification_type.capitalize()} 验证混淆矩阵')

        # 设置刻度标签
        plt.xticks(np.arange(3), ['原始', '蒸馏', '伪装'])
        plt.yticks(np.arange(2), ['通过', '未通过'])

        # 添加颜色条
        plt.colorbar(label='模型数量')

        # 保存图表
        plt.tight_layout()
        plt.savefig(os.path.join(images_dir, f"confusion_matrix_{verification_type}.png"), dpi=300)
        plt.close()

    print(f"Visualizations saved to {images_dir}")


def create_excel_report(original_results, distilled_results, disguised_results, verification_results, analysis_result):
    """
    创建Excel报告

    Args:
        original_results: 原始模型训练结果
        distilled_results: 蒸馏模型训练结果
        disguised_results: 伪装模型训练结果
        verification_results: 验证结果
        analysis_result: 分析结果

    Returns:
        Excel报告路径
    """
    # 创建工作簿
    wb = Workbook()

    # 创建模型训练结果表
    ws_models = wb.active
    ws_models.title = "模型训练结果"

    # 添加标题
    headers = ["模型类型", "模型ID", "准确率", "其他参数", "证明路径", "模型路径"]
    for col, header in enumerate(headers, 1):
        cell = ws_models.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 添加数据 - 原始模型
    row = 2
    for result in original_results:
        ws_models.cell(row=row, column=1, value="原始模型")
        ws_models.cell(row=row, column=2, value=result["model_id"])
        ws_models.cell(row=row, column=3, value=result["accuracy"])
        ws_models.cell(row=row, column=4, value="无")
        ws_models.cell(row=row, column=5, value=result.get("proof_path", "无"))
        ws_models.cell(row=row, column=6, value=result["model_path"])
        row += 1

    # 添加数据 - 蒸馏模型
    for result in distilled_results:
        ws_models.cell(row=row, column=1, value="蒸馏模型")
        ws_models.cell(row=row, column=2, value=result["model_id"])
        ws_models.cell(row=row, column=3, value=result["accuracy"])
        ws_models.cell(row=row, column=4, value=f"温度={result['temperature']}")
        ws_models.cell(row=row, column=5, value=result.get("proof_path", "无"))
        ws_models.cell(row=row, column=6, value=result["model_path"])
        row += 1

    # 添加数据 - 伪装模型
    for result in disguised_results:
        ws_models.cell(row=row, column=1, value="伪装模型")
        ws_models.cell(row=row, column=2, value=result["model_id"])
        ws_models.cell(row=row, column=3, value=result["accuracy"])
        ws_models.cell(row=row, column=4, value=f"伪装率={result['disguise_rate']}")
        ws_models.cell(row=row, column=5, value=result.get("proof_path", "无"))
        ws_models.cell(row=row, column=6, value=result["model_path"])
        row += 1

    # 自动调整列宽
    for col in ws_models.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws_models.column_dimensions[column].width = adjusted_width

    # 创建验证结果表
    ws_verification = wb.create_sheet(title="验证结果")

    # 添加标题
    headers = ["模型类型", "模型ID", "验证类型", "通过验证", "结果路径"]
    for col, header in enumerate(headers, 1):
        cell = ws_verification.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 添加数据
    for row, result in enumerate(verification_results, 2):
        model_type_map = {
            "original": "原始模型",
            "distilled": "蒸馏模型",
            "disguised": "伪装模型"
        }
        verification_type_map = {
            "whitebox": "白盒验证",
            "blackbox": "黑盒验证",
            "combined": "组合验证"
        }

        ws_verification.cell(row=row, column=1, value=model_type_map.get(result["model_type"], result["model_type"]))
        ws_verification.cell(row=row, column=2, value=result["model_id"])
        ws_verification.cell(row=row, column=3,
                             value=verification_type_map.get(result["verification_type"], result["verification_type"]))

        # 为通过/失败添加颜色
        cell = ws_verification.cell(row=row, column=4, value="通过" if result["passed"] else "失败")
        if result["passed"]:
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
        else:
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色

        ws_verification.cell(row=row, column=5, value=result["result_path"])

    # 自动调整列宽
    for col in ws_verification.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws_verification.column_dimensions[column].width = adjusted_width

    # 创建分析结果表
    ws_analysis = wb.create_sheet(title="分析结果")

    # 添加通过率数据
    ws_analysis.cell(row=1, column=1, value="通过率 (%)")
    ws_analysis.cell(row=1, column=1).font = Font(bold=True)

    ws_analysis.cell(row=2, column=1, value="验证方法")
    ws_analysis.cell(row=2, column=2, value="原始模型")
    ws_analysis.cell(row=2, column=3, value="蒸馏模型")
    ws_analysis.cell(row=2, column=4, value="伪装模型")

    verification_types = ["whitebox", "blackbox", "combined"]
    verification_names = ["白盒验证", "黑盒验证", "组合验证"]

    for i, (vt, vn) in enumerate(zip(verification_types, verification_names)):
        row = i + 3
        ws_analysis.cell(row=row, column=1, value=vn)

        original_key = f"original_{vt}"
        distilled_key = f"distilled_{vt}"
        disguised_key = f"disguised_{vt}"

        ws_analysis.cell(row=row, column=2, value=analysis_result["pass_rates"].get(original_key, 0))
        ws_analysis.cell(row=row, column=3, value=analysis_result["pass_rates"].get(distilled_key, 0))
        ws_analysis.cell(row=row, column=4, value=analysis_result["pass_rates"].get(disguised_key, 0))

    # 添加空行
    ws_analysis.cell(row=7, column=1, value="")

    # 添加检测准确率数据
    ws_analysis.cell(row=8, column=1, value="检测准确率 (%)")
    ws_analysis.cell(row=8, column=1).font = Font(bold=True)

    ws_analysis.cell(row=9, column=1, value="验证方法")
    ws_analysis.cell(row=9, column=2, value="准确率")

    for i, (vt, vn) in enumerate(zip(verification_types, verification_names)):
        row = i + 10
        ws_analysis.cell(row=row, column=1, value=vn)
        ws_analysis.cell(row=row, column=2, value=analysis_result["detection_accuracy"].get(vt, 0))

    # 自动调整列宽
    for col in ws_analysis.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws_analysis.column_dimensions[column].width = adjusted_width

    # 保存工作簿
    report_path = os.path.join(BASE_PATH, "模型验证报告.xlsx")
    wb.save(report_path)

    print(f"Excel report saved to {report_path}")

    return report_path


def run_validity_experiment(resume=False, load_models=False, model_dir=None):
    """
    运行验证有效性实验

    Args:
        resume: 是否恢复训练
        load_models: 是否加载已有模型
        model_dir: 模型目录路径

    Returns:
        Excel报告路径
    """
    print("\n===== 开始验证有效性实验 =====\n")

    # 创建实验目录
    setup_experiment_directories()

    # 加载数据集
    trainloader, testloader = load_dataset(use_small=USE_SMALL_DATASET)

    # 训练或加载原始模型
    if load_models and model_dir:
        # 加载已有模型
        print("\n==== 加载原始模型 ====\n")
        original_models, original_results = load_original_models(model_dir, num_models=3)
    else:
        # 训练新模型
        print("\n==== 训练原始模型 ====\n")
        original_models, original_results = train_original_models(
            trainloader, testloader, num_models=3, resume=resume
        )

    # 选择一个原始模型作为教师模型
    teacher_model = original_models[0]

    # 训练蒸馏模型
    print("\n==== 训练蒸馏模型 ====\n")
    distilled_models, distilled_results = train_distilled_models(
        teacher_model, trainloader, testloader, num_models=3, resume=resume
    )

    # 训练伪装模型
    print("\n==== 训练伪装模型 ====\n")
    disguised_models, disguised_results = train_disguised_models(
        teacher_model, trainloader, testloader, num_models=2
    )

    # 验证所有模型
    print("\n==== 验证模型 ====\n")
    verification_results = verify_models(
        original_models,
        distilled_models,
        disguised_models,
        original_results,
        distilled_results,
        disguised_results,
        testloader
    )

    # 分析验证结果
    print("\n==== 分析验证结果 ====\n")
    analysis_result = analyze_verification_results(
        verification_results,
        original_results,
        distilled_results,
        disguised_results
    )

    # 输出简要分析
    print("\n==== 验证结果摘要 ====\n")
    print("通过率:")
    for key, rate in analysis_result["pass_rates"].items():
        print(f"  {key}: {rate:.1f}%")

    print("\n检测准确率:")
    for key, accuracy in analysis_result["detection_accuracy"].items():
        print(f"  {key}: {accuracy:.1f}%")

    # 创建可视化
    print("\n==== 创建可视化 ====\n")
    create_verification_visualizations(analysis_result)

    # 创建Excel报告
    print("\n==== 创建Excel报告 ====\n")
    report_path = create_excel_report(
        original_results,
        distilled_results,
        disguised_results,
        verification_results,
        analysis_result
    )

    print("\n===== 验证有效性实验完成 =====\n")

    return report_path


if __name__ == "__main__":
    # 测试运行
    report_path = run_validity_experiment()
    print(f"\n报告已保存至: {report_path}")